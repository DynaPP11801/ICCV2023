# YOLOv5 🚀 by Ultralytics, GPL-3.0 license
"""
Validate a trained YOLOv5 detection model on a detection dataset

Usage:
    $ python val.py --weights yolov5s.pt --data coco128.yaml --img 640

Usage - formats:
    $ python val.py --weights yolov5s.pt                 # PyTorch
                              yolov5s.torchscript        # TorchScript
                              yolov5s.onnx               # ONNX Runtime or OpenCV DNN with --dnn
                              yolov5s_openvino_model     # OpenVINO
                              yolov5s.engine             # TensorRT
                              yolov5s.mlmodel            # CoreML (macOS-only)
                              yolov5s_saved_model        # TensorFlow SavedModel
                              yolov5s.pb                 # TensorFlow GraphDef
                              yolov5s.tflite             # TensorFlow Lite
                              yolov5s_edgetpu.tflite     # TensorFlow Edge TPU
                              yolov5s_paddle_model       # PaddlePaddle
"""

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
import torch
from tqdm import tqdm

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from models.common import DetectMultiBackend
from utils.callbacks import Callbacks
from utils.dataloaders import create_dataloader
from utils.general import (LOGGER, TQDM_BAR_FORMAT, Profile, check_dataset, check_img_size, check_requirements,
                           check_yaml, coco80_to_coco91_class, colorstr, increment_path, non_max_suppression,
                           print_args, scale_boxes, xywh2xyxy, xyxy2xywh)
from utils.metrics import ConfusionMatrix, ap_per_class, box_iou, box_center
from utils.plots import output_to_target, plot_images, plot_val_study
from utils.torch_utils import select_device, smart_inference_mode

from canvas_DynaPP import *
from openpyxl import Workbook
from openpyxl import load_workbook
import torchvision



def save_one_txt(predn, save_conf, shape, file):
    # Save one txt result
    gn = torch.tensor(shape)[[1, 0, 1, 0]]  # normalization gain whwh
    for *xyxy, conf, cls in predn.tolist():
        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
        with open(file, 'a') as f:
            f.write(('%g ' * len(line)).rstrip() % line + '\n')


def save_one_json(predn, jdict, path, class_map):
    # Save one JSON result {"image_id": 42, "category_id": 18, "bbox": [258.15, 41.29, 348.26, 243.78], "score": 0.236}
    image_id = int(path.stem) if path.stem.isnumeric() else path.stem
    box = xyxy2xywh(predn[:, :4])  # xywh
    box[:, :2] -= box[:, 2:] / 2  # xy center to top-left corner
    for p, b in zip(predn.tolist(), box.tolist()):
        jdict.append({
            'image_id': image_id,
            'category_id': class_map[int(p[5])],
            'bbox': [round(x, 3) for x in b],
            'score': round(p[4], 5)})


def process_batch(detections, labels, iouv):
    """
    Return correct prediction matrix
    Arguments:
        detections (array[N, 6]), x1, y1, x2, y2, conf, class
        labels (array[M, 5]), class, x1, y1, x2, y2
    Returns:
        correct (array[N, 10]), for 10 IoU levels
    """
    correct = np.zeros((detections.shape[0], iouv.shape[0])).astype(bool)
    iou = box_iou(labels[:, 1:], detections[:, :4])
    correct_class = labels[:, 0:1] == detections[:, 5]
    for i in range(len(iouv)):
        x = torch.where((iou >= iouv[i]) & correct_class)  # IoU > threshold and classes match
        if x[0].shape[0]:
            matches = torch.cat((torch.stack(x, 1), iou[x[0], x[1]][:, None]), 1).cpu().numpy()  # [label, detect, iou]
            if x[0].shape[0] > 1:
                matches = matches[matches[:, 2].argsort()[::-1]]
                matches = matches[np.unique(matches[:, 1], return_index=True)[1]]
                # matches = matches[matches[:, 2].argsort()[::-1]]
                matches = matches[np.unique(matches[:, 0], return_index=True)[1]]
            correct[matches[:, 1].astype(int), i] = True
    return torch.tensor(correct, dtype=torch.bool, device=iouv.device)

def AUAIR_change_cls(cls):
    change = torch.cuda.LongTensor([0,0,5,1,3,2,5,5,6,4])
    return change[cls.type('torch.cuda.LongTensor')]


def UAVDT_change_cls(cls):
    change = torch.cuda.LongTensor([-1,-1,-1,0,-1,1,-1,-1,2,-1])
    return change[cls.type('torch.cuda.LongTensor')]


@smart_inference_mode()
def run(
        data,
        weights=None,  # model.pt path(s)
        batch_size=32,  # batch size
        imgsz=640,  # inference size (pixels)
        conf_thres=0.001,  # confidence threshold
        iou_thres=0.6,  # NMS IoU threshold
        max_det=300,  # maximum detections per image
        task='val',  # train, val, test, speed or study
        device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        workers=8,  # max dataloader workers (per RANK in DDP mode)
        single_cls=False,  # treat as single-class dataset
        augment=False,  # augmented inference
        verbose=False,  # verbose output
        save_txt=False,  # save results to *.txt
        save_hybrid=False,  # save label+prediction hybrid results to *.txt
        save_conf=False,  # save confidences in --save-txt labels
        save_json=False,  # save a COCO-JSON results file
        project=ROOT / 'runs/val',  # save to project/name
        name='exp',  # save to project/name
        exist_ok=False,  # existing project/name ok, do not increment
        half=True,  # use FP16 half-precision inference
        dnn=False,  # use OpenCV DNN for ONNX inference

        whether_pack=False,
        dataset_name='None',
        duration=5,
        background=0.3,
        saved_frames=100,
        minimum=33,
    
        model=None,
        dataloader=None,
        save_dir=Path(''),
        plots=True,
        callbacks=Callbacks(),
        compute_loss=None,
):
    # Initialize/load model and set device
    training = model is not None
    if training:  # called by train.py
        device, pt, jit, engine = next(model.parameters()).device, True, False, False  # get model device, PyTorch model
        half &= device.type != 'cpu'  # half precision only supported on CUDA
        model.half() if half else model.float()
    else:  # called directly
        device = select_device(device, batch_size=batch_size)

        # Directories
        save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
        (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

        # Load model
        model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
        stride, pt, jit, engine = model.stride, model.pt, model.jit, model.engine
#         imgsz = check_img_size(imgsz, s=stride)  # check image size
        half = model.fp16  # FP16 supported on limited backends with CUDA
        if engine:
            batch_size = model.batch_size
        else:
            device = model.device
            if not (pt or jit):
                batch_size = 1  # export.py models default to batch-size 1
                LOGGER.info(f'Forcing --batch-size 1 square inference (1,3,{imgsz},{imgsz}) for non-PyTorch models')

        # Data
        data = check_dataset(data)  # check

    # Configure
    model.eval()
    cuda = device.type != 'cpu'
    is_coco = isinstance(data.get('val'), str) and data['val'].endswith(f'coco{os.sep}val2017.txt')  # COCO dataset
    nc = 1 if single_cls else int(data['nc'])  # number of classes
    iouv = torch.linspace(0.5, 0.95, 10, device=device)  # iou vector for mAP@0.5:0.95
    niou = iouv.numel()

    # Dataloader
    if not training:
        if pt and not single_cls:  # check --weights are trained on --data
            ncm = model.model.nc
#             assert ncm == nc, f'{weights} ({ncm} classes) trained on different --data than what you passed ({nc} ' \
#                               f'classes). Pass correct combination of --weights and --data that are trained together.'
        model.warmup(imgsz=(1 if pt else batch_size, 3, imgsz, imgsz))  # warmup
        pad, rect = (0.0, False) if task == 'speed' else (0.5, pt)  # square inference for benchmarks
        task = task if task in ('train', 'val', 'test') else 'val'  # path to train/val/test images
        
        if not training:
            pad, rect = (0.0, False)
            
        if whether_pack == True:
            rect = True
            pad = 0
      
        
        dataloader = create_dataloader(data[task],
                                       imgsz,
                                       batch_size,
                                       stride,
                                       single_cls,
                                       pad=pad,
                                       rect=rect,
                                       workers=workers,
                                       prefix=colorstr(f'{task}: '))[0]

    seen = 0
    confusion_matrix = ConfusionMatrix(nc=nc)
#     names = model.names if hasattr(model, 'names') else model.module.names  # get class names
#     print(names, 'model names\n')
    names = data['names']
#     print(names, 'data names\n')
#     input()
    if isinstance(names, (list, tuple)):  # old format
        names = dict(enumerate(names))
    class_map = coco80_to_coco91_class() if is_coco else list(range(1000))
    s = ('%22s' + '%11s' * 6) % ('Class', 'Images', 'Instances', 'P', 'R', 'mAP50', 'mAP50-95')
    tp, fp, p, r, f1, mp, mr, map50, ap50, map = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    
#     t0, t1,t2,t3, t_pack, pack_t1, pack_t2
    dt = Profile(), Profile(), Profile(), Profile(), Profile(), Profile()
    packed_img_h, packed_img_w = [] , []
    all_packed_img = []
    all_anchor_img = []  
    key_frame = 0
    prev_out = None
    duration = duration
    if imgsz==960:  
        div_shap = 24
    elif imgsz==2560: 
        div_shap = 32
    else:
        div_shap = 1
    all_img_shape = np.zeros([div_shap,div_shap])
    div_sz = imgsz//div_shap
    if not training:
        f = save_dir / 'packed_frames'
        os.mkdir(f)
    
    
#     dt = Profile(), Profile(), Profile()  # profiling times
    loss = torch.zeros(3, device=device)
    jdict, stats, ap, ap_class = [], [], [], []
    callbacks.run('on_val_start')
    pbar = tqdm(dataloader, desc=s, bar_format=TQDM_BAR_FORMAT)  # progress bar
    for batch_i, (im, targets, paths, shapes) in enumerate(pbar):
        callbacks.run('on_val_batch_start')
        with dt[0]:
            if cuda:
                im = im.to(device, non_blocking=True)
                targets = targets.to(device)
            im = im.half() if half else im.float()  # uint8 to fp16/32
            im /= 255  # 0 - 255 to 0.0 - 1.0
            nb, _, height, width = im.shape  # batch size, channels, height, width

        # Inference
        with torch.no_grad():
            if whether_pack == False or key_frame ==0:
                with dt[1]:
                    preds, train_out = model(im) if compute_loss else (model(im, augment=augment), None)
                
                all_img_shape[(im.size(2)+div_sz//2)//div_sz-1, (im.size(3)+div_sz//2)//div_sz-1] += 1
                # Loss
                if compute_loss:
                    loss += compute_loss(train_out, targets)[1]  # box, obj, cls

                # NMS
                targets[:, 2:] *= torch.tensor((width, height, width, height), device=device)  # to pixels
                lb = [targets[targets[:, 0] == i, 1:] for i in range(nb)] if save_hybrid else []  # for autolabelling


                with dt[2]:
                    preds = non_max_suppression(preds,
                                                conf_thres,
                                                iou_thres,
                                                labels=lb,
                                                multi_label=True,
                                                agnostic=single_cls,
                                                max_det=max_det)                    
                    
                    
                pack = False
                if whether_pack == True:
                    where_old_image= None
                    where_new_image = None
                    all_anchor_img.append(im.size(2)*im.size(3))

                    if len(preds[0] != 0):
                        key_frame +=1
                        old_out = [[]]
                        old_out[0] = preds[0].clone().type(torch.cuda.LongTensor)
                        prev_out = preds[0][:,:4]
                    else:
                        key_frame = 0  
                                 
            else:
                pack = True
                with dt[5]:
                    packed_img, where_old_image, where_new_image, groups_groups = canvas(im, preds[0][:,:4], where_old_image, where_new_image, prev_out, background, minimum)
                    
                with dt[4]:
                    if key_frame ==1:    
                        box_minmax = [old_out[0][groups,:4] for groups in groups_groups]
                        old_box_flow_remember = []
                        for boxs in box_minmax:
                            if len(boxs) !=0:
                                old_box_flow_remember.append(torch.cat([torch.min(boxs[:,:2], dim=0, keepdim=True)[0],torch.max(boxs[:,2:], dim=0, keepdim=True)[0]], dim=1).type(torch.cuda.LongTensor))
                            else:
                                old_box_flow_remember.append(None)                    
                    
                packed_img = F.pad(packed_img, (0,max(0,64-packed_img.size(3)),0,max(0,64- packed_img.size(2))), "constant", 0)
                all_packed_img.append(packed_img.size(2)*packed_img.size(3))
                
                with dt[1]:
                    preds, train_out = model(packed_img) if compute_loss else (model(packed_img, augment=augment), None)  
                    
                a = packed_img.size(2)
                b = packed_img.size(3)
                if (a+div_sz//2)//div_sz-1>=div_shap or (b+div_sz//2)//div_sz-1>=div_shap:
                    import math
                    temp_a = math.sqrt(a*b)
                    temp_b = a*b/temp_a
                    
                    a, b = int(temp_a), int(temp_b)
                    
                    
                all_img_shape[(a+div_sz//2)//div_sz-1, (b+div_sz//2)//div_sz-1] += 1
                
                # Loss
                if compute_loss:
                    loss += compute_loss(train_out, targets)[1]  # box, obj, cls

                # NMS
                targets[:, 2:] *= torch.tensor((width, height, width, height), device=device)  # to pixels
                lb = [targets[targets[:, 0] == i, 1:] for i in range(nb)] if save_hybrid else []  # for autolabelling


                with dt[2]:
                    preds = non_max_suppression(preds,
                                                conf_thres,
                                                iou_thres,
                                                labels=lb,
                                                multi_label=True,
                                                agnostic=single_cls,
                                                max_det=max_det)                          
                    
                if len(preds[0]) != 0: 
                    if where_new_image == None:
                        with dt[3]:
                            preds[0][:,:2] += where_old_image[:,:2]
                            preds[0][:,2:4] += where_old_image[:,:2]
                            argmax = torch.zeros(len(preds[0])).type(torch.cuda.LongTensor)

                    else:
                        with dt[3]:
                            pack_box_iou = box_center(preds[0][:,:4], where_new_image)
                            argmax = torch.argmax(pack_box_iou, dim=1)

                            check_for_edge = preds[0][:,:4] - where_new_image[argmax,:4]
                            edge_thres = max(im.size(3), im.size(2))//100
                            index_for_chips_edge = (check_for_edge[:,0:1]<edge_thres)*1+(check_for_edge[:,1:2]<edge_thres)*1+(check_for_edge[:,2:3]>-edge_thres)*1+(check_for_edge[:,3:4]>-edge_thres)*1<=0

                            preds[0][:,2:4] -= where_new_image[argmax,:2]
                            preds[0][:,:2] -= where_new_image[argmax,:2]
                            preds[0][:,:2] += where_old_image[argmax,:2]
                            preds[0][:,2:4] += where_old_image[argmax,:2]
                            
                            index_for_img_edge = 1*(preds[0][:,0:1]<edge_thres)+1*(preds[0][:,1:2]<edge_thres)+1*(preds[0][:,2:3]>im.size(3)-edge_thres)+1*(preds[0][:,3:4]>im.size(2)-edge_thres)>0

                            index_for_edge = index_for_chips_edge+index_for_img_edge>0

                            index_for_edge = index_for_edge.nonzero(as_tuple=True)[0]
                            preds = [torch.index_select(preds[0], 0, index_for_edge)]

                        with dt[2]:
                            preds = [preds[0][torchvision.ops.nms(preds[0][:,:4], preds[0][:,4], iou_thres)]]                  
                        if len(preds[0]) != 0: 
                            pack_box_iou = box_center(preds[0][:,:4], where_old_image)
                            argmax = torch.argmax(pack_box_iou, dim=1)
 

                            with dt[4]:
                                box_minmax = [preds[0][argmax==k,:4].clone().type(torch.cuda.LongTensor) for k in range(len(where_old_image))]
                                box_flow_remember = []
                                for boxs in box_minmax:
                                    if len(boxs) !=0:
                                        box_flow_remember.append(torch.cat([torch.min(boxs[:,:2], dim=0, keepdim=True)[0],torch.max(boxs[:,2:], dim=0, keepdim=True)[0]], dim=1))
                                    else:
                                        box_flow_remember.append(None)
                                limit = max(im.size(3), im.size(2))//20
                                for index, (box_flow, old_box_flow) in enumerate(zip(box_flow_remember, old_box_flow_remember)):
                                    if box_flow is None and old_box_flow is not None: 
                                        box_flow_remember[index] = old_box_flow
                                    if box_flow is None or old_box_flow is None:
                                        continue
                                    x_min_flow, x_max_flow, y_min_flow, y_max_flow = box_flow[0,0]-old_box_flow[0,0], box_flow[0,2]-old_box_flow[0,2], box_flow[0,1]-old_box_flow[0,1], box_flow[0,3]-old_box_flow[0,3]
                                    x_min_flow, x_max_flow, y_min_flow, y_max_flow = x_min_flow.item(), x_max_flow.item(), y_min_flow.item(), y_max_flow.item()
                                    if x_min_flow*x_max_flow>0 and abs(x_min_flow)<limit and abs(x_max_flow)<limit:
                                        some_to_add = [x_min_flow,x_max_flow][np.argmin([abs(x_min_flow),abs(x_max_flow)])]
                                        if some_to_add<0:
                                            if where_old_image[index,0]+some_to_add < 0:
                                                some_to_add = -where_old_image[index,0]
                                        else:
                                            if where_old_image[index,2]+some_to_add > im.size(3):
                                                some_to_add = im.size(3)-where_old_image[index,2]
                                        where_old_image[index,[0,2]] += some_to_add
                                    if y_min_flow*y_max_flow>0 and abs(y_min_flow)<limit and abs(y_max_flow)<limit:
                                        some_to_add = [y_min_flow,y_max_flow][np.argmin([abs(y_min_flow),abs(y_max_flow)])]
                                        if some_to_add<0:
                                            if where_old_image[index,1]+some_to_add < 0:
                                                some_to_add = -where_old_image[index,1]
                                        else:
                                            if where_old_image[index,3]+some_to_add > im.size(2):
                                                some_to_add = im.size(2)-where_old_image[index,3]
                                        where_old_image[index,[1,3]] += some_to_add

                                old_box_flow_remember = box_flow_remember

                    prev_out = preds[0][:,:4]

                if len(preds[0]) != 0:
                    key_frame +=1             
                else:
                    key_frame = 0 

                if key_frame ==duration:
                    key_frame=0                             
                            
                    




    
        # Metrics
        for si, pred in enumerate(preds):
            labels = targets[targets[:, 0] == si, 1:]
            nl, npr = labels.shape[0], pred.shape[0]  # number of labels, predictions
            path, shape = Path(paths[si]), shapes[si][0]
            correct = torch.zeros(npr, niou, dtype=torch.bool, device=device)  # init
            seen += 1

            if npr == 0:
                if nl:
                    stats.append((correct, *torch.zeros((2, 0), device=device), labels[:, 0]))
                    if plots:
                        confusion_matrix.process_batch(detections=None, labels=labels[:, 0])
                continue

            # Predictions
            if single_cls:
                pred[:, 5] = 0
            if dataset_name =='AUAIR':
                pred[:, 5] = AUAIR_change_cls(pred[:, 5])
            if dataset_name =='UAVDT':
                pred[:, 5] = UAVDT_change_cls(pred[:, 5])
            
            pred, preds[si] = pred[pred[:, 5]!=-1], preds[si][pred[:, 5]!=-1]
            
            predn = pred.clone()
            scale_boxes(im[si].shape[1:], predn[:, :4], shape, shapes[si][1])  # native-space pred

            # Evaluate
            if nl:
                tbox = xywh2xyxy(labels[:, 1:5])  # target boxes
                scale_boxes(im[si].shape[1:], tbox, shape, shapes[si][1])  # native-space labels
                labelsn = torch.cat((labels[:, 0:1], tbox), 1)  # native-space labels
                correct = process_batch(predn, labelsn, iouv)
                if plots:
                    confusion_matrix.process_batch(predn, labelsn)
            stats.append((correct, pred[:, 4], pred[:, 5], labels[:, 0]))  # (correct, conf, pcls, tcls)

            # Save/log
            if save_txt:
                save_one_txt(predn, save_conf, shape, file=save_dir / 'labels' / f'{path.stem}.txt')
            if save_json:
                save_one_json(predn, jdict, path, class_map)  # append to COCO-JSON dictionary
            callbacks.run('on_val_image_end', pred, predn, path, names, im[si])

        # Plot images
        if plots and batch_i < saved_frames:
            n = '0'*(4-len(str(batch_i)))+str(batch_i)
            f = save_dir / f'test_batch{n}_labels.jpg'  # labels
            plot_images(im, targets, paths, f, names)  # labels
            f = save_dir / f'test_batch{n}_pred.jpg'  # predictions
            plot_images(im, output_to_target(preds), paths, f, names)  # pred
            if pack == True:
                f = save_dir / f'packed_frames/pack_image{n}.jpg'  # packed
                plot_images(packed_img, output_to_target([preds[0]*0]), paths, f, names)  # packed

        callbacks.run('on_val_batch_end', batch_i, im, targets, paths, shapes, preds)

    # Compute metrics
    stats = [torch.cat(x, 0).cpu().numpy() for x in zip(*stats)]  # to numpy
    if len(stats) and stats[0].any():
        tp, fp, p, r, f1, ap, ap_class = ap_per_class(*stats, plot=plots, save_dir=save_dir, names=names)
        ap50, ap = ap[:, 0], ap.mean(1)  # AP@0.5, AP@0.5:0.95
        mp, mr, map50, map = p.mean(), r.mean(), ap50.mean(), ap.mean()
    nt = np.bincount(stats[3].astype(int), minlength=nc)  # number of targets per class

    # Print results
    pf = '%22s' + '%11i' * 2 + '%11.3g' * 4  # print format
    LOGGER.info(pf % ('all', seen, nt.sum(), mp, mr, map50, map))
    if nt.sum() == 0:
        LOGGER.warning(f'WARNING ⚠️ no labels found in {task} set, can not compute metrics without labels')

    # Print results per class
    if (verbose or (nc < 50 and not training)) and nc > 1 and len(stats):
        for i, c in enumerate(ap_class):
            LOGGER.info(pf % (names[c], seen, nt[c], p[i], r[i], ap50[i], ap[i]))

    # Print speeds
    t = tuple(x.t / seen * 1E3 for x in dt)  # speeds per image
    
    
    
    if not training:
        shape = (batch_size, 3, imgsz, imgsz)
        LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS, %.1fms Back_to_origianl, %.1fms Pad_Track, %.1fms Packing_time per image at shape {shape}' % t)

    # Saved in excel
    if not training:
        if whether_pack == False:
            load_wb = load_workbook(f'excel_result/{dataset_name}_results_baseline.xlsx', data_only = True)
            all_size_img= np.mean(all_anchor_img)
        else:
            load_wb = load_workbook(f'excel_result/{dataset_name}_results_DynaPP.xlsx', data_only = True)
            all_size_img = np.mean(all_packed_img+all_anchor_img)
        all_size_img = np.sqrt(all_size_img)
        load_ws = load_wb['Sheet']
        all_values = []
        all_img_shape = all_img_shape.flatten()
        all_img_shape= all_img_shape.tolist()

        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        wr = [data['val'].split('/')[-2],t[5],t[1],t[2],t[3],t[4],t[5]+t[1]+t[2]+t[3]+t[4],map50, map, all_size_img]
        wr.extend(all_img_shape)
        all_values.append(wr)

        write_wb =Workbook()
        write_ws = write_wb.active
        for i in range(len(all_values)):
            write_ws.append(list(all_values[i]))
        if whether_pack == False:
            write_wb.save(f'excel_result/{dataset_name}_results_baseline.xlsx')
        else:
            write_wb.save(f'excel_result/{dataset_name}_results_DynaPP.xlsx')
            
        
        
    # Plots
    if plots:
        confusion_matrix.plot(save_dir=save_dir, names=list(names.values()))
        callbacks.run('on_val_end', nt, tp, fp, p, r, f1, ap, ap50, ap_class, confusion_matrix)

    # Save JSON
    if save_json and len(jdict):
        w = Path(weights[0] if isinstance(weights, list) else weights).stem if weights is not None else ''  # weights
        anno_json = str(Path('../datasets/coco/annotations/instances_val2017.json'))  # annotations
        pred_json = str(save_dir / f'{w}_predictions.json')  # predictions
        LOGGER.info(f'\nEvaluating pycocotools mAP... saving {pred_json}...')
        with open(pred_json, 'w') as f:
            json.dump(jdict, f)

        try:  # https://github.com/cocodataset/cocoapi/blob/master/PythonAPI/pycocoEvalDemo.ipynb
            check_requirements('pycocotools>=2.0.6')
            from pycocotools.coco import COCO
            from pycocotools.cocoeval import COCOeval

            anno = COCO(anno_json)  # init annotations api
            pred = anno.loadRes(pred_json)  # init predictions api
            eval = COCOeval(anno, pred, 'bbox')
            if is_coco:
                eval.params.imgIds = [int(Path(x).stem) for x in dataloader.dataset.im_files]  # image IDs to evaluate
            eval.evaluate()
            eval.accumulate()
            eval.summarize()
            map, map50 = eval.stats[:2]  # update results (mAP@0.5:0.95, mAP@0.5)
        except Exception as e:
            LOGGER.info(f'pycocotools unable to run: {e}')

    # Return results
    model.float()  # for training
    if not training:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    maps = np.zeros(nc) + map
    for i, c in enumerate(ap_class):
        maps[c] = ap[i]
    return (mp, mr, map50, map, *(loss.cpu() / len(dataloader)).tolist()), maps, t


def parse_opt():
    parser = argparse.ArgumentParser()
    parser.add_argument('--data', type=str, default=ROOT / 'data/coco128.yaml', help='dataset.yaml path')
    parser.add_argument('--weights', nargs='+', type=str, default=ROOT / 'yolov5s.pt', help='model path(s)')
    parser.add_argument('--batch-size', type=int, default=32, help='batch size')
    parser.add_argument('--imgsz', '--img', '--img-size', type=int, default=640, help='inference size (pixels)')
    parser.add_argument('--conf-thres', type=float, default=0.001, help='confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.6, help='NMS IoU threshold')
    parser.add_argument('--max-det', type=int, default=300, help='maximum detections per image')
    parser.add_argument('--task', default='val', help='train, val, test, speed or study')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--workers', type=int, default=8, help='max dataloader workers (per RANK in DDP mode)')
    parser.add_argument('--single-cls', action='store_true', help='treat as single-class dataset')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--verbose', action='store_true', help='report mAP by class')
    parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
    parser.add_argument('--save-hybrid', action='store_true', help='save label+prediction hybrid results to *.txt')
    parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
    parser.add_argument('--save-json', action='store_true', help='save a COCO-JSON results file')
    parser.add_argument('--project', default=ROOT / 'runs/val', help='save to project/name')
    parser.add_argument('--name', default='exp', help='save to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    parser.add_argument('--half', action='store_true', help='use FP16 half-precision inference')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')

    parser.add_argument('--whether_pack', action='store_true', help='Use packing')
    parser.add_argument('--dataset_name', type=str, default='None', help='dataset name')
    parser.add_argument('--duration', type=int, default=5, help='key frame duration length')
    parser.add_argument('--background', type=float, default=0.3, help='background d%')
    parser.add_argument('--saved_frames', type=int, default=100, help='number of the saved frames')
    parser.add_argument('--minimum', type=int, default=33, help='minimum length of image (reverse/2 %)')

    opt = parser.parse_args()
    opt.data = check_yaml(opt.data)  # check YAML
    opt.save_json |= opt.data.endswith('coco.yaml')
    opt.save_txt |= opt.save_hybrid
    print_args(vars(opt))
    return opt


def main(opt):
    check_requirements(exclude=('tensorboard', 'thop'))

    if opt.task in ('train', 'val', 'test'):  # run normally
        if opt.conf_thres > 0.001:  # https://github.com/ultralytics/yolov5/issues/1466
            LOGGER.info(f'WARNING ⚠️ confidence threshold {opt.conf_thres} > 0.001 produces invalid results')
        if opt.save_hybrid:
            LOGGER.info('WARNING ⚠️ --save-hybrid will return high mAP from hybrid labels, not from predictions alone')
        run(**vars(opt))

    else:
        weights = opt.weights if isinstance(opt.weights, list) else [opt.weights]
        opt.half = torch.cuda.is_available() and opt.device != 'cpu'  # FP16 for fastest results
        if opt.task == 'speed':  # speed benchmarks
            # python val.py --task speed --data coco.yaml --batch 1 --weights yolov5n.pt yolov5s.pt...
            opt.conf_thres, opt.iou_thres, opt.save_json = 0.25, 0.45, False
            for opt.weights in weights:
                run(**vars(opt), plots=False)

        elif opt.task == 'study':  # speed vs mAP benchmarks
            # python val.py --task study --data coco.yaml --iou 0.7 --weights yolov5n.pt yolov5s.pt...
            for opt.weights in weights:
                f = f'study_{Path(opt.data).stem}_{Path(opt.weights).stem}.txt'  # filename to save to
                x, y = list(range(256, 1536 + 128, 128)), []  # x axis (image sizes), y axis
                for opt.imgsz in x:  # img-size
                    LOGGER.info(f'\nRunning {f} --imgsz {opt.imgsz}...')
                    r, _, t = run(**vars(opt), plots=False)
                    y.append(r + t)  # results and times
                np.savetxt(f, y, fmt='%10.4g')  # save
            subprocess.run(['zip', '-r', 'study.zip', 'study_*.txt'])
            plot_val_study(x=x)  # plot
        else:
            raise NotImplementedError(f'--task {opt.task} not in ("train", "val", "test", "speed", "study")')


if __name__ == '__main__':
    opt = parse_opt()
    main(opt)